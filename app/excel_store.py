from __future__ import annotations

import threading
from datetime import datetime
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.config import DATA_DIR, EXCEL_PATH

COLUMNS = [
    "Fecha",
    "Nombre",
    "Teléfono / WhatsApp",
    "Email",
    "Ciudad / localidad",
    "Tipo de propiedad",
    "Servicio de interés",
    "Mensaje",
    "Origen",
]

LOCK = threading.Lock()
try:
    TZ = ZoneInfo("America/Argentina/Buenos_Aires")
except Exception:
    TZ = None

HEADER_FILL = PatternFill("solid", fgColor="0B1C24")
HEADER_FONT = Font(name="Calibri", bold=True, color="00E5C3", size=11)
CELL_FONT = Font(name="Calibri", color="1A1A1A", size=11)
ALT_FILL = PatternFill("solid", fgColor="F3FAF8")
THIN = Border(
    left=Side(style="thin", color="D7E4E0"),
    right=Side(style="thin", color="D7E4E0"),
    top=Side(style="thin", color="D7E4E0"),
    bottom=Side(style="thin", color="D7E4E0"),
)


def _ensure_workbook() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    if EXCEL_PATH.exists():
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Consultas"
    ws.append(COLUMNS)

    for col, _ in enumerate(COLUMNS, start=1):
        cell = ws.cell(1, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    ws.row_dimensions[1].height = 22
    _autosize(ws)
    wb.save(EXCEL_PATH)


def _autosize(ws) -> None:
    widths = [22, 28, 22, 32, 22, 22, 24, 50, 16]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width


def save_consulta(payload: dict) -> None:
    with LOCK:
        _ensure_workbook()
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
        row = [
            datetime.now(TZ).strftime("%Y-%m-%d %H:%M") if TZ else datetime.now().strftime("%Y-%m-%d %H:%M"),
            payload["nombre"],
            payload["telefono"],
            payload["email"],
            payload["ciudad"],
            payload["tipo_propiedad"],
            payload["servicio"],
            payload.get("mensaje") or "",
            payload.get("origen") or "landing",
        ]
        ws.append(row)
        row_idx = ws.max_row
        for col in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row_idx, col)
            cell.font = CELL_FONT
            cell.border = THIN
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if row_idx % 2 == 0:
                cell.fill = ALT_FILL
        ws.row_dimensions[row_idx].height = 32
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{row_idx}"
        wb.save(EXCEL_PATH)
